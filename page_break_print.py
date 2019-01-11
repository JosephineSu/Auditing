from openpyxl.workbook import workbook
from openpyxl import load_workbook
from openpyxl.worksheet.pagebreak import Break
from openpyxl.styles.borders import Border,Side
import pandas as pd


def page_break_print(fp):
    # wb = Workbook()
    # ws = wb.active
    wb = load_workbook(fp)
    table = pd.read_excel(fp)
    ws = wb.active
    # print(type(table))
    count = 0
    row_number = 1
    flag = True
    # header = ['指标', '12月', '01月', '02月', '03月', '04月', '05月', '06月', '07月', '08月', '09月', '10月', '11月', '合计']
    for index,row in table.iterrows():
        if flag == True:
            flag = False
        else:
            # print(row['指标'], type(row['指标']))
            if len(str(row['指标']).split(' ')[0]) != 21:
                count = count + 1
                continue
            row_number += count  # the row that you want to insert page break
            row_number += 1
            # print(count,row_number)
            page_break = Break(id=row_number)  # create Break obj
            ws.page_breaks.append(page_break)  # insert page break
            count = 0
    index += 2
    rows = ws['A1':'N{}'.format(index)]
    rows = list(rows)
    # set_border(wb)
    thin_border = Border(left=Side(style='dashed'),
                         right=Side(style='dashed'),
                         top=Side(style='dashed'),
                         bottom=Side(style='dashed'))
    for x,cells in enumerate(rows):
        for y,cell in enumerate(cells):
            ws.cell(row=x+1,column=y+1).border = thin_border
    wb.save(fp)

# # 打开csv文件 返回DataFrame对象
# def read_excel(path):
#     with open(path, 'r') as f:
#         file = pd.read_excel(io=0)
#     return file


if __name__ == '__main__':
    fp = 'D:\研一\项目\Auditing\所有乡镇所有居委会台账结果.xlsx'
    # table = pd.read_excel(fp)
    page_break_print(fp)